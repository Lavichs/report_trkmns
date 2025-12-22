import xlrd
import httpx
import shutil
import pandas as pd
from openpyxl import load_workbook
from camoufox.sync_api import Camoufox
from browserforge.fingerprints import Screen


titles = [
    "Здание Детский стационар, ул. Комсомольская д. 200",
    "Здание Консультативно-диагностический центр, ул. Кобозева д. 25, к. А",
    "Здание Поликлиника №1, ул. Рыбаковская д.3, помещение 2",
    "Здание Поликлиника №2, ул. Пойменная д. 23, к. А",
    "Здание Поликлиника №3, ул. Алтайская д. 2",
    "Здание Поликлиника №4, ул. Туркестанская д. 43",
    "Здание Отделение медицинской реабилитации, ул. Кобозева д.12",
]
date_beg = "09.12.2025"
date_end = "15.12.2025"

TEMPLATE_EXCEL = "Template.xlsx"
DEST_EXCEL = f"Отчет по ИС ЛО с {date_beg} по {date_end}.xlsx"


def add_data_to_dest_excel(data: list, row: int):
    wb_dest = load_workbook(DEST_EXCEL)
    ws_dest = wb_dest.active
    i = 2

    for item in data:
        ws_dest.cell(row=row, column=i).value = item
        i += 1

    ws_dest.cell(row=10, column=1).value = f"{date_beg} - {date_end}"

    wb_dest.save(DEST_EXCEL)


def download_orders():
    # constrains = Screen(max_width=1280, max_height=720)
    # with Camoufox(screen=constrains) as browser:
    with Camoufox(headless=True) as browser:
        context = browser.new_context(
            accept_downloads=True, viewport={"width": 1280, "height": 720}
        )
        page = context.new_page()

        # page = browser.new_page()
        page.goto("https://ecp56.is-mis.ru/?c=portal&m=udp")
        page.wait_for_timeout(5000)

        login = page.locator("#promed-login")
        page.locator("#promed-login").fill("shuriginIA")
        page.locator("#promed-password").fill("#$&password#$??")
        page.locator("#auth_submit").click()

        page.wait_for_timeout(15000)
        page.get_by_text("Отчеты").click()
        page.wait_for_timeout(3000)

        el = (
            page.locator(".x-tree-root-node")
            .locator(":scope > :last-child")
            .locator("> *")
            .nth(0)
            .locator("> *")
            .nth(1)
        )
        el.click()
        page.wait_for_timeout(1000)
        page.get_by_text("Список случаев лечения").click()
        page.wait_for_timeout(2000)

        # ------------ Перешли на страницу отчета ------------
        # ------------ В цикле качаем нужные отчеты ------------
        page.locator('[id="rpt.engine.paramBegDate"]').fill(date_beg)
        page.locator('[id="rpt.engine.paramEndDate"]').fill(date_end)
        page.locator('[id="rpt.engine.paramEvnType"] + *').click()
        page.get_by_text("КВС").click()
        for title in titles:
            # ------------ Редактируем параметры отчета ------------

            # page.locator('#rpt.engine.paramBegDate').fill("09.12.2025")
            # page.locator('[id="rpt.engine.paramEvnType"]').fill("КВС")
            page.locator('[id="rpt.engine.paramLpuBuilding"] + *').click()
            page.locator('[id="rpt.engine.paramLpuBuilding"]').type(title)
            # page.locator('div.x-panel-body').click()

            page.locator(
                "div.x-combo-list-item.x-combo-selected", has_text=title
            ).click()

            # ------------ Параметры отчета изменены. Скачиваем отчет ------------

            page.locator(
                '[matomo_event_id="win_swReportEndUserWindow_tbr_undefined"]'
            ).click()
            page.get_by_text("Формат XLSX").click()
            page.get_by_text("Сформировать отчет").click()

            with page.expect_download(timeout=40000) as download_info:
                page.get_by_text("Сформировать отчет").click()

            download = download_info.value
            download.save_as(f"uploads/{title}.xlsx")

            # ------------ Отчет скачан ------------
            page.wait_for_timeout(5000)  # таймаут пока прогрузиться страница


def get_max_stap_first():
    data = []
    for title in titles:
        wb = load_workbook(f"uploads/{title}.xlsx", data_only=True)
        ws = wb.active

        value = ws.cell(row=ws.max_row - 1, column=1).value
        if ws.cell(row=ws.max_row - 1, column=3).value == "3":
            value = 0
        data.append(value)

    add_data_to_dest_excel(data, 2)


def islo_parse_1():
    url_islo = "http://172.30.149.11:8282/new_islo/web/admin/hospitalization-management"
    login = "MALAHOVA_DS"
    password = "AS6P2BH0"

    # with Camoufox(screen=constrains) as browser:
    # constrains = Screen(max_width=1280, max_height=720)
    with Camoufox(headless=True) as browser:
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        # page = browser.new_page()
        page.goto(url_islo)
        # page.wait_for_timeout(5000)

        page.locator('[id="loginform-username"]').type(login)
        page.locator('[id="loginform-password"]').type(password)
        page.get_by_text("Войти").click()

        page.wait_for_timeout(7000)

        lctr = page.locator("*", has_text="Выбор медицинской организации")

        if lctr.count() > 0:
            page.locator(
                '[id="select2-update-organization-form-select-container"]'
            ).click()

            # page.get_by_text('се ГАУЗ "ДГКБ"').click()
            # page.locator('select2-search__field').type('ОРЕНБУРГ ГАУЗ "ДГКБ"')
            # page.locator('select2-search__field').press("Enter")
            page.locator(".select2-results__option", has_text="ОРЕНБУРГ").nth(1).click()
            page.locator(".btn", has_text="Сохранить").click()

            page.wait_for_timeout(5000)

        try:
            page.locator('[id="treatmenthospital-bbdate"]').fill(
                f"{date_beg} - {date_end}"
            )
            page.locator(".form-check-label", has_text="Все").click()
            page.locator("#w2-button").click()

            with page.expect_download(timeout=40000) as download_info:
                page.locator("#w0-xlsx").click()

                page.locator(".btn", has_text="Ok").click()

            download = download_info.value
            download.save_as(f"uploads/islo.xlsx")

        except Exception as e:
            print("expt")


def get_data_from_islo():
    wb = load_workbook(f"uploads/islo.xlsx", data_only=True)
    ws = wb.active

    data = {}

    for i in range(1, ws.max_row + 1):
        data[ws.cell(row=i, column=4).value] = (
            data.get(ws.cell(row=i, column=4).value, 0) + 1
        )

    data = {
        "Стационар круглосуточный": data.get("Педиатрическое отделение №1", 0)
        + data.get("Педиатрическое отделение №2", 0)
        + data.get("Отделение анестезиологии и реанимации", 0)
        + data.get("Отделение патологии детей раннего возраста", 0),
        "Дневные стационары": data.get("ДС при АПУ Педиатрическое отделение (КДЦ)", 0)
        + data.get("ДС при АПУ Хирургическое отд. (КДЦ)", 0)
        + data.get("ДС при АПУ Офтальмологическое отделение (КДЦ)", 0),
        "Пол-ка 1": data.get("СНД Поликлиника №1", 0),
        "Пол-ка 2": data.get("СНД Поликлиника №2", 0),
        "Пол-ка 3": data.get("СНД Поликлиника №3", 0),
        "Пол-ка 4": data.get("СНД Поликлиника №4", 0),
        "Реабилитация": data.get("Медицинская реабилитация (ДС)", 0),
    }

    values = []
    for k, v in data.items():
        values.append(v)

    add_data_to_dest_excel(values, 3)


def islo_parse_2():
    url_islo = "http://172.30.149.11:8282/new_islo/web"
    login = "MALAHOVA_DS"
    password = "AS6P2BH0"

    # with Camoufox(headless=True) as browser:
    constrains = Screen(max_width=1280, max_height=720)
    with Camoufox(screen=constrains) as browser:
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()

        # page = browser.new_page()
        page.goto(url_islo)
        # page.wait_for_timeout(5000)

        page.locator('[id="loginform-username"]').type(login)
        page.locator('[id="loginform-password"]').type(password)
        page.get_by_text("Войти").click()

        page.wait_for_timeout(7000)

        lctr = page.locator("*", has_text="Выбор медицинской организации")

        if lctr.count() > 0:
            page.locator(
                '[id="select2-update-organization-form-select-container"]'
            ).click()

            # page.get_by_text('се ГАУЗ "ДГКБ"').click()
            # page.locator('select2-search__field').type('ОРЕНБУРГ ГАУЗ "ДГКБ"')
            # page.locator('select2-search__field').press("Enter")
            page.locator(".select2-results__option", has_text="ОРЕНБУРГ").nth(1).click()
            page.locator(".btn", has_text="Сохранить").click()

            page.wait_for_timeout(5000)

        page.goto("http://172.30.149.11:8282/OE/appointment/allappointments")

        page.wait_for_timeout(5000)

        page.locator("#start_date").fill("9 Дек 2025")
        page.locator("#end_date").fill("15 Дек 2025")
        page.locator("#searchallapp").click()
        page.wait_for_timeout(3000)

        with page.expect_download(timeout=40000) as download_info:
            page.locator("#allappexel").click()

        download = download_info.value
        download.save_as(f"uploads/islo_2.xlsx")

        # page.locator('#allappexel').click()


def get_data_from_islo_2():
    df = pd.read_excel("uploads/islo_2.ods", engine="odf")
    data = {}
    data_on_100 = {}

    for i in range(0, len(df)):
        if "Действующий" in df.iloc[i, 2] or "Отменён" in df.iloc[i, 2]:
            continue

        data[df.iloc[i, 3]] = data.get(df.iloc[i, 3], 0) + 1

        if "100.0%" in df.iloc[i, 12]:
            data_on_100[df.iloc[i, 3]] = data_on_100.get(df.iloc[i, 3], 0) + 1

    data = {
        "Стационар круглосуточный": data.get("Педиатрическое отделение №1", 0)
        + data.get("Педиатрическое отделение №2", 0)
        + data.get("Отделение анестезиологии и реанимации", 0)
        + data.get("Отделение патологии детей раннего возраста", 0),
        "Дневные стационары": data.get("ДС при АПУ Педиатрическое отделение (КДЦ)", 0)
        + data.get("ДС при АПУ Хирургическое отд. (КДЦ)", 0)
        + data.get("ДС при АПУ Офтальмологическое отделение (КДЦ)", 0),
        "Пол-ка 1": data.get("СНД Поликлиника №1", 0),
        "Пол-ка 2": data.get("СНД Поликлиника №2", 0),
        "Пол-ка 3": data.get("СНД Поликлиника №3", 0),
        "Пол-ка 4": data.get("СНД Поликлиника №4", 0),
        "Реабилитация": data.get("Медицинская реабилитация (ДС)", 0),
    }
    data_on_100 = {
        "Стационар круглосуточный": data_on_100.get("Педиатрическое отделение №1", 0)
        + data_on_100.get("Педиатрическое отделение №2", 0)
        + data_on_100.get("Отделение анестезиологии и реанимации", 0)
        + data_on_100.get("Отделение патологии детей раннего возраста", 0),
        "Дневные стационары": data_on_100.get(
            "ДС при АПУ Педиатрическое отделение (КДЦ)", 0
        )
        + data_on_100.get("ДС при АПУ Хирургическое отд. (КДЦ)", 0)
        + data_on_100.get("ДС при АПУ Офтальмологическое отделение (КДЦ)", 0),
        "Пол-ка 1": data_on_100.get("СНД Поликлиника №1", 0),
        "Пол-ка 2": data_on_100.get("СНД Поликлиника №2", 0),
        "Пол-ка 3": data_on_100.get("СНД Поликлиника №3", 0),
        "Пол-ка 4": data_on_100.get("СНД Поликлиника №4", 0),
        "Реабилитация": data_on_100.get("Медицинская реабилитация (ДС)", 0),
    }

    values = []
    values_2 = []
    for k, v in data.items():
        values.append(v)
        values_2.append(data_on_100.get(k))

    add_data_to_dest_excel(values, 6)
    add_data_to_dest_excel(values_2, 7)


def main():
    shutil.copy2(TEMPLATE_EXCEL, DEST_EXCEL)
    # line 1
    # download_orders()
    get_max_stap_first()
    # line 2
    # islo_parse_1()
    get_data_from_islo()
    # line 3
    # islo_parse_2()
    get_data_from_islo_2()


if __name__ == "__main__":
    main()


# from camoufox.sync_api import Camoufox

# with Camoufox() as browser:
#     page = browser.new_page()
#     page.goto("https://ya.ru")
#     page.wait_for_timeout(5000)
