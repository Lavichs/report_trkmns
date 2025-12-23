import json
import base64
import shutil
import locale
import logging
import asyncio
import datetime
import subprocess
import pandas as pd
from pathlib import Path
from datetime import timedelta
from openpyxl import load_workbook
from fastapi import FastAPI, BackgroundTasks
from fastapi.responses import FileResponse
# from camoufox.sync_api import Camoufox
from camoufox.async_api import AsyncCamoufox
# from playwright.async_api import async_playwright
from browserforge.fingerprints import Screen
# from faststream.rabbit.fastapi import RabbitRouter
from faststream.rabbit import RabbitBroker
from contextlib import asynccontextmanager

from config import consts


# router = RabbitRouter()
broker = RabbitBroker(consts.RABBIT_URL)

logger = logging.getLogger("uvicorn")

locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")
UPLOADS_PATH = Path("uploads")
TEMPLATE_EXCEL = "Template.xlsx"


# Определить функцию-потребителя
@broker.subscriber("incoming")
async def on_incoming(msg: str):
    # Обработать сообщение
    # logger.info(f"Received message: {msg}")
    # logger.info(f"Start proccessing")

    logger.info(msg)
    msg = json.loads(msg)
    # await broker.publish(msg, queue="response")

    try:
        if not Path(consts.DEST_EXCEL_PATH).exists():
            shutil.copy2(TEMPLATE_EXCEL, consts.DEST_EXCEL_PATH)

            await getting_data_from_ECP()
            await getting_hospitalisation_data()
            await getting_data_about_appointment_list()

            proccess_ECP_data()
            proccess_islo_hospitalisation_data()
            proccess_islo_appointment_list_data()

        with open(consts.DEST_EXCEL_PATH, "rb") as f:
            file_bytes = f.read()

        file_base64 = base64.b64encode(file_bytes).decode('utf-8')

        # попытка выполнить скрипт
        msg['text'] = 'Скрипт отработал успешно'
        msg['content'] = file_base64
        msg['filename'] = consts.DEST_EXCEL_TITLE
        await broker.publish(json.dumps(msg), queue="response")
    except Exception as e:
        logger.error(e)
        msg['text'] = 'Ошибка во время выполнения'
        await broker.publish(json.dumps(msg), queue="response")
    
    # print(f"Received message: {msg}")
    # Если нужно, вернуть ответ
    # return Prediction(msg)


@asynccontextmanager
async def lifespan(app: FastAPI):
    await broker.start()
    yield
    await broker.close()


app = FastAPI(lifespan=lifespan)


@app.post("/report")
def get_report(background_tasks: BackgroundTasks):
    # background_tasks.add_task(getting_data_from_ECP)
    # await getting_hospitalisation_data()
    # await getting_data_about_appointment_list()

    if not Path(consts.DEST_EXCEL_PATH).exists():
        shutil.copy2(TEMPLATE_EXCEL, consts.DEST_EXCEL_PATH)

        # getting data
        # await getting_data_from_ECP()
        # await getting_hospitalisation_data()
        # await getting_data_about_appointment_list()

        # Proccess data
        # proccess_ECP_data()
        # proccess_islo_hospitalisation_data()
        # proccess_islo_appointment_list_data()

    return {'data': "Tasks start"}

    # return FileResponse(
    #     path=consts.DEST_EXCEL_PATH,
    #     filename=consts.DEST_EXCEL_TITLE,
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    # )


def add_data_to_dest_excel(data: list, row: int):
    wb_dest = load_workbook(consts.DEST_EXCEL_PATH)
    ws_dest = wb_dest.active
    i = 2

    for item in data:
        ws_dest.cell(row=row, column=i).value = item
        i += 1

    ws_dest.cell(row=10, column=1).value = f"{consts.WEEK_AGO} - {consts.YESTERDAY}"

    wb_dest.save(consts.DEST_EXCEL_PATH)


async def getting_data_from_ECP():
    # constrains = Screen(max_width=1280, max_height=720)
    # with Camoufox(screen=constrains) as browser:
    # async with async_playwright() as p:
    # with Camoufox(headless=True) as browser:
    async with AsyncCamoufox(headless=True) as browser:
        # context = browser.new_context(
        #     accept_downloads=True, viewport={"width": 1280, "height": 720}
        # )
        # page = context.new_page()
        page = await browser.new_page()
        logger.info("Start Camoufox")

        # page = browser.new_page()
        await page.goto("https://ecp56.is-mis.ru/?c=portal&m=udp")
        await page.wait_for_timeout(5000)
        logger.info("Load ECP")

        login = page.locator("#promed-login")
        await page.locator("#promed-login").fill(consts.ECP_USER)
        await page.locator("#promed-password").fill(consts.ECP_PASS)
        await page.locator("#auth_submit").click()

        await page.wait_for_timeout(15000)
        await page.get_by_text("Отчеты").click()
        await page.wait_for_timeout(3000)
        logger.info("Load reports page")

        el = (
            page.locator(".x-tree-root-node")
            .locator(":scope > :last-child")
            .locator("> *")
            .nth(0)
            .locator("> *")
            .nth(1)
        )
        await el.click()
        await page.wait_for_timeout(1000)
        await page.get_by_text("Список случаев лечения").click()
        await page.wait_for_timeout(15000)

        # ------------ Перешли на страницу отчета ------------
        # ------------ В цикле качаем нужные отчеты ------------
        await page.locator('[id="rpt.engine.paramBegDate"]').fill(consts.WEEK_AGO)
        await page.locator('[id="rpt.engine.paramEndDate"]').fill(consts.YESTERDAY)
        await page.locator('[id="rpt.engine.paramEvnType"] + *').click()
        await page.get_by_text("КВС").click()
        for title in consts.TITLES:
            # ------------ Редактируем параметры отчета ------------

            await page.locator('[id="rpt.engine.paramLpuBuilding"] + *').click()
            await page.locator('[id="rpt.engine.paramLpuBuilding"]').type(title)

            await page.locator(
                "div.x-combo-list-item.x-combo-selected", has_text=title
            ).click()

            # ------------ Параметры отчета изменены. Скачиваем отчет ------------

            await page.locator(
                '[matomo_event_id="win_swReportEndUserWindow_tbr_undefined"]'
            ).click()
            await page.get_by_text("Формат XLSX").click()
            await page.get_by_text("Сформировать отчет").click()

            async with page.expect_download(timeout=90000) as download_info:
                await page.get_by_text("Сформировать отчет").click()

                download = await download_info.value
                await download.save_as(UPLOADS_PATH.joinpath(f"{title}.xlsx"))

            # ------------ Отчет скачан ------------
            logger.info(f"Load report [{consts.TITLES.index(title)}]")
            await page.wait_for_timeout(5000)  # таймаут пока прогрузиться страница
        logger.info("End loading")


async def getting_hospitalisation_data():
    # with Camoufox(screen=constrains) as browser:
    # constrains = Screen(max_width=1280, max_height=720)
    # with Camoufox(headless=True) as browser:
    async with AsyncCamoufox(headless=True) as browser:
        # context = browser.new_context(accept_downloads=True)
        # page = context.new_page()
        page = await browser.new_page()
        await page.goto(
            "http://172.30.149.11:8282/new_islo/web/admin/hospitalization-management"
        )
        logger.info("Start Camoufox")

        await page.locator('[id="loginform-username"]').type(consts.ISLO_USER)
        await page.locator('[id="loginform-password"]').type(consts.ISLO_PASS)
        await page.get_by_text("Войти").click()
        await page.wait_for_timeout(7000)
        lctr = page.locator("*", has_text="Выбор медицинской организации")
        if await lctr.count() > 0:
            await page.locator(
                '[id="select2-update-organization-form-select-container"]'
            ).click()
            await page.locator(".select2-results__option", has_text="ОРЕНБУРГ").nth(1).click()
            await page.locator(".btn", has_text="Сохранить").click()
            await page.wait_for_timeout(5000)
        
        logger.info("Authorisation ISLO")

        try:
            await page.locator('[id="treatmenthospital-bbdate"]').fill(
                f"{consts.WEEK_AGO} - {consts.YESTERDAY}"
            )
            await page.locator(".form-check-label", has_text="Все").click()
            await page.locator("#w2-button").click()

            logger.info("Start load report")

            async with page.expect_download(timeout=40000) as download_info:
                await page.locator("#w0-xlsx").click()
                await page.locator(".btn", has_text="Ok").click()

                download = await download_info.value
                await download.save_as(UPLOADS_PATH.joinpath("islo_hospitalisation.xlsx"))

            logger.info("End loading")

        except Exception as e:
            print("expt")


async def getting_data_about_appointment_list():
    # with Camoufox(headless=True) as browser:
    # constrains = Screen(max_width=1280, max_height=720)
    # with Camoufox(screen=constrains) as browser:
    async with AsyncCamoufox(headless=True) as browser:
        # context = browser.new_context(accept_downloads=True)
        # page = context.new_page()
        page = await browser.new_page()
        await page.goto("http://172.30.149.11:8282/new_islo/web")
        
        logger.info("Start Camoufox")

        await page.locator('[id="loginform-username"]').type(consts.ISLO_USER)
        await page.locator('[id="loginform-password"]').type(consts.ISLO_PASS)
        await page.get_by_text("Войти").click()
        await page.wait_for_timeout(7000)
        lctr = page.locator("*", has_text="Выбор медицинской организации")
        if await lctr.count() > 0:
            await page.locator(
                '[id="select2-update-organization-form-select-container"]'
            ).click()
            await page.locator(".select2-results__option", has_text="ОРЕНБУРГ").nth(1).click()
            await page.locator(".btn", has_text="Сохранить").click()
            await page.wait_for_timeout(5000)

        logger.info("Authorisation ISLO")

        await page.goto("http://172.30.149.11:8282/OE/appointment/allappointments")
        await page.wait_for_timeout(5000)
        date_beg = (
            datetime.datetime.strptime(consts.WEEK_AGO, "%d.%m.%Y")
            .strftime("%-d %b %Y")
            .title()
        )
        date_end = (
            datetime.datetime.strptime(consts.WEEK_AGO, "%d.%m.%Y")
            .strftime("%-d %b %Y")
            .title()
        )
        await page.locator("#start_date").fill(date_beg)
        await page.locator("#end_date").fill(date_end)
        await page.locator("#searchallapp").click()
        await page.wait_for_timeout(3000)
        
        logger.info("Start load report")

        async with page.expect_download(timeout=40000) as download_info:
            await page.locator("#allappexel").click()

            download = await download_info.value
            await download.save_as(UPLOADS_PATH.joinpath("islo_appointment_list.xlsx"))

        logger.info("End loading")


def proccess_ECP_data():
    data = []
    for title in consts.TITLES:
        wb = load_workbook(UPLOADS_PATH.joinpath(f"{title}.xlsx"), data_only=True)
        ws = wb.active

        value = ws.cell(row=ws.max_row - 1, column=1).value
        if ws.cell(row=ws.max_row - 1, column=3).value == "3":
            value = 0
        data.append(value)

    add_data_to_dest_excel(data, 2)


def proccess_islo_hospitalisation_data():
    wb = load_workbook(
        UPLOADS_PATH.joinpath("islo_hospitalisation.xlsx"), data_only=True
    )
    ws = wb.active
    data = {}
    for i in range(1, ws.max_row + 1):
        data[ws.cell(row=i, column=4).value] = (
            data.get(ws.cell(row=i, column=4).value, 0) + 1
        )
    data = {
        "Стационар круглосуточный": data.get("Педиатрическое отделение №1", 0)
        + data.get("Педиатрическое отделение №2", 0)
        + data.get("Отделение анестезиологии и реанимации", 0)
        + data.get("Отделение патологии детей раннего возраста", 0),
        "Дневные стационары": data.get("ДС при АПУ Педиатрическое отделение (КДЦ)", 0)
        + data.get("ДС при АПУ Хирургическое отд. (КДЦ)", 0)
        + data.get("ДС при АПУ Офтальмологическое отделение (КДЦ)", 0),
        "Пол-ка 1": data.get("СНД Поликлиника №1", 0),
        "Пол-ка 2": data.get("СНД Поликлиника №2", 0),
        "Пол-ка 3": data.get("СНД Поликлиника №3", 0),
        "Пол-ка 4": data.get("СНД Поликлиника №4", 0),
        "Реабилитация": data.get("Медицинская реабилитация (ДС)", 0),
    }
    values = []
    for k, v in data.items():
        values.append(v)

    add_data_to_dest_excel(values, 3)


def proccess_islo_appointment_list_data():
    subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "ods",
            "--outdir",
            "./uploads",
            "./uploads/islo_appointment_list.xlsx",
        ]
    )
    df = pd.read_excel(UPLOADS_PATH.joinpath("islo_appointment_list.ods"), engine="odf")
    data = {}
    data_on_100 = {}
    for i in range(0, len(df)):
        if "Действующий" in df.iloc[i, 2] or "Отменён" in df.iloc[i, 2]:
            continue
        data[df.iloc[i, 3]] = data.get(df.iloc[i, 3], 0) + 1
        if "100.0%" in df.iloc[i, 12]:
            data_on_100[df.iloc[i, 3]] = data_on_100.get(df.iloc[i, 3], 0) + 1

    data = {
        "Стационар круглосуточный": data.get("Педиатрическое отделение №1", 0)
        + data.get("Педиатрическое отделение №2", 0)
        + data.get("Отделение анестезиологии и реанимации", 0)
        + data.get("Отделение патологии детей раннего возраста", 0),
        "Дневные стационары": data.get("ДС при АПУ Педиатрическое отделение (КДЦ)", 0)
        + data.get("ДС при АПУ Хирургическое отд. (КДЦ)", 0)
        + data.get("ДС при АПУ Офтальмологическое отделение (КДЦ)", 0),
        "Пол-ка 1": data.get("СНД Поликлиника №1", 0),
        "Пол-ка 2": data.get("СНД Поликлиника №2", 0),
        "Пол-ка 3": data.get("СНД Поликлиника №3", 0),
        "Пол-ка 4": data.get("СНД Поликлиника №4", 0),
        "Реабилитация": data.get("Медицинская реабилитация (ДС)", 0),
    }
    data_on_100 = {
        "Стационар круглосуточный": data_on_100.get("Педиатрическое отделение №1", 0)
        + data_on_100.get("Педиатрическое отделение №2", 0)
        + data_on_100.get("Отделение анестезиологии и реанимации", 0)
        + data_on_100.get("Отделение патологии детей раннего возраста", 0),
        "Дневные стационары": data_on_100.get(
            "ДС при АПУ Педиатрическое отделение (КДЦ)", 0
        )
        + data_on_100.get("ДС при АПУ Хирургическое отд. (КДЦ)", 0)
        + data_on_100.get("ДС при АПУ Офтальмологическое отделение (КДЦ)", 0),
        "Пол-ка 1": data_on_100.get("СНД Поликлиника №1", 0),
        "Пол-ка 2": data_on_100.get("СНД Поликлиника №2", 0),
        "Пол-ка 3": data_on_100.get("СНД Поликлиника №3", 0),
        "Пол-ка 4": data_on_100.get("СНД Поликлиника №4", 0),
        "Реабилитация": data_on_100.get("Медицинская реабилитация (ДС)", 0),
    }
    values = []
    values_2 = []
    for k, v in data.items():
        values.append(v)
        values_2.append(data_on_100.get(k))

    add_data_to_dest_excel(values, 6)
    add_data_to_dest_excel(values_2, 7)
